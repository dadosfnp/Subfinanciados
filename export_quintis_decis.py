"""
Script one-off: exporta media de receita per capita (rc_24_pc) agregada
por quintil e decil para um arquivo Excel.

Uso: python export_quintis_decis.py
"""
import os
import sys
import re
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from django.db.models import Avg, Count, Min, Max
from home.models import Municipio
from openpyxl import Workbook


def quantile_sort_key(label):
    """Ordena '1º quintil', '2º quintil' etc. pela ordem numerica."""
    match = re.match(r'(\d+)', label or '')
    return int(match.group(1)) if match else 999


def aggregate_by(field):
    """Retorna [(label, n_municipios, media_rc, min_rc, max_rc), ...] ordenado."""
    qs = (
        Municipio.objects
        .exclude(**{f'{field}__isnull': True})
        .exclude(rc_24_pc__isnull=True)
        .values(field)
        .annotate(
            n=Count('id'),
            media=Avg('rc_24_pc'),
            minimo=Min('rc_24_pc'),
            maximo=Max('rc_24_pc'),
        )
    )
    rows = [(r[field], r['n'], r['media'], r['minimo'], r['maximo']) for r in qs]
    rows.sort(key=lambda r: quantile_sort_key(r[0]))
    return rows


def write_sheet(ws, title, rows):
    ws.title = title
    headers = [title.rstrip('s'), 'Nº de municípios', 'Média RC per capita (R$)',
               'Mín. RC per capita (R$)', 'Máx. RC per capita (R$)']
    ws.append(headers)
    for label, n, media, minimo, maximo in rows:
        ws.append([label, n, round(media, 2), round(minimo, 2), round(maximo, 2)])
    # Largura das colunas
    widths = [18, 18, 26, 26, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    # Negrito no header
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)


def main():
    quintis = aggregate_by('quintil24')
    decis = aggregate_by('decil24')

    wb = Workbook()
    write_sheet(wb.active, 'Quintis', quintis)
    write_sheet(wb.create_sheet('Decis'), 'Decis', decis)

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'quintis_decis_media_per_capita.xlsx')
    wb.save(out_path)
    print(f'Arquivo salvo: {out_path}')
    print()
    print('=== Quintis ===')
    for r in quintis:
        print(f'  {r[0]:>15} | n={r[1]:>4} | média=R$ {r[2]:,.2f}'.replace(',', '.'))
    print()
    print('=== Decis ===')
    for r in decis:
        print(f'  {r[0]:>15} | n={r[1]:>4} | média=R$ {r[2]:,.2f}'.replace(',', '.'))


if __name__ == '__main__':
    main()
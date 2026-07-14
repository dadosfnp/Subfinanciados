"""
Script one-off: exporta dados do IFEM para a lista de "Cidades Intermediadoras
para o Desenvolvimento Regional" (Regiões Imediatas — RI) em Excel.

Para cada município da lista busca no banco por (uf, name_muni) com matching
tolerante a acentos/caixa. Os que não bateram entram com a coluna "Status"
marcada como "Não encontrado".
"""
import os
import sys
import unicodedata
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from home.models import Municipio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# -- Lista extraída do PDF "Lista cidades intermediadoras" -------------------
# Estrutura: (UF, nome da RI, [municípios])
GRUPOS = [
    ('AC', 'Cruzeiro do Sul', [
        'Cruzeiro do Sul', 'Mâncio Lima', 'Marechal Thaumaturgo',
        'Porto Walter', 'Rodrigues Alves',
    ]),
    ('AL', 'Porto Calvo - São Luís do Quitunde', [
        'Campestre', 'Colônia Leopoldina', 'Jacuípe', 'Japaratinga',
        'Jundiá', 'Maragogi', 'Matriz de Camaragibe', 'Novo Lino',
        'Passo de Camaragibe', 'Porto Calvo', 'Porto de Pedras',
        'São Luís do Quitunde', 'São Miguel dos Milagres',
    ]),
    ('AP', 'Oiapoque', [
        'Amapá', 'Calçoene', 'Cutias', 'Oiapoque', 'Pracuúba',
        'Tartarugalzinho',
    ]),
    ('AM', 'Tefé', [
        'Alvarães', 'Carauari', 'Fonte Boa', 'Japurá', 'Juruá',
        'Jutaí', 'Maraã', 'Tefé', 'Uarini',
    ]),
    ('BA', 'Xique-Xique / Barra', [
        'Barra', 'Brotas de Macaúbas', 'Buritirama', 'Gentio do Ouro',
        'Ibotirama', 'Ipupiara', 'Morpará', 'Muquém de São Francisco',
        'Oliveira dos Brejinhos', 'Xique-Xique',
    ]),
    ('CE', 'Itapipoca', [
        'Amontada', 'Itapipoca', 'Miraíma', 'Trairi', 'Tururu',
        'Umirim', 'Uruburetama',
    ]),
    ('ES', 'São Mateus', [
        'Boa Esperança', 'Conceição da Barra', 'Jaguaré', 'Montanha',
        'Mucurici', 'Pedro Canário', 'Pinheiros', 'Ponto Belo', 'São Mateus',
    ]),
    ('GO', 'Posse / Campos Belos', [
        'Alvorada do Norte', 'Buritinópolis', 'Campos Belos',
        'Damianópolis', 'Divinópolis de Goiás', 'Guarani de Goiás',
        'Iaciara', 'Mambaí', 'Monte Alegre de Goiás', 'Nova Roma',
        'Posse', 'São Domingos', 'Simolândia', "Sítio d'Abadia",
    ]),
    ('MA', 'Santa Inês', [
        'Alto Alegre do Pindaré', 'Araguanã', 'Bela Vista do Maranhão',
        'Bom Jardim', 'Governador Newton Bello', 'Igarapé do Meio',
        'Monção', 'Nova Olinda do Maranhão', 'Pindaré-Mirim', 'Pio XII',
        'Santa Inês', 'Santa Luzia', 'São João do Carú', 'Tufilândia',
        'Zé Doca',
    ]),
    ('MT', 'Cáceres', [
        'Cáceres', 'Curvelândia', "Lambari D'Oeste", 'Rio Branco',
        'Salto do Céu',
    ]),
    ('MS', 'Corumbá', [
        'Corumbá', 'Ladário',
    ]),
    ('MG', 'Araçuaí', [
        'Araçuaí', 'Berilo', 'Coronel Murta', 'Francisco Badaró',
        'Itinga', 'Jenipapo de Minas', 'José Gonçalves de Minas',
        'Virgem da Lapa',
    ]),
    ('PA', 'Breves', [
        'Afuá', 'Anajás', 'Bagre', 'Breves', 'Chaves', 'Curralinho',
        'Gurupá', 'Melgaço', 'Portel', 'São Sebastião da Boa Vista',
    ]),
    ('PB', 'Cajazeiras', [
        'Bom Jesus', 'Bonito de Santa Fé', 'Cachoeira dos Índios',
        'Cajazeiras', 'Carrapateira', 'Monte Horebe',
        'Poço de José de Moura', 'São João do Rio do Peixe',
        'Santa Helena', 'São José de Piranhas', 'Serra Grande', 'Triunfo',
    ]),
    ('PR', 'Laranjeiras do Sul / Quedas do Iguaçu', [
        'Espigão Alto do Iguaçu', 'Laranjeiras do Sul', 'Marquinho',
        'Nova Laranjeiras', 'Porto Barreiro', 'Quedas do Iguaçu',
        'Rio Bonito do Iguaçu', 'Virmond',
    ]),
    ('PE', 'Serra Talhada', [
        'Betânia', 'Calumbi', 'Carnaubeira da Penha', 'Flores',
        'Floresta', 'Jatobá', 'Mirandiba', 'Petrolândia',
        'Santa Cruz da Baixa Verde', 'São José do Belmonte',
        'Serra Talhada', 'Tacaratu', 'Triunfo',
    ]),
    ('PI', 'Parnaíba', [
        'Bom Princípio do Piauí', 'Buriti dos Lopes',
        'Cajueiro da Praia', 'Caraúbas do Piauí', 'Caxingó',
        'Cocal', 'Cocal dos Alves', 'Ilha Grande', 'Luís Correia',
        'Murici dos Portelas', 'Parnaíba',
    ]),
    ('RJ', 'Rio Bonito', [
        'Cachoeiras de Macacu', 'Rio Bonito', 'Silva Jardim',
    ]),
    ('RN', 'Mossoró', [
        'Apodi', 'Areia Branca', 'Augusto Severo', 'Baraúna', 'Caraúbas',
        'Felipe Guerra', 'Governador Dix-Sept Rosado', 'Grossos', 'Itaú',
        'Janduís', 'Messias Targino', 'Mossoró', 'Rodolfo Fernandes',
        'Tibau', 'Serra do Mel', 'Severiano Melo', 'Upanema',
    ]),
    ('RS', 'Uruguaiana', [
        'Alegrete', 'Barra do Quaraí', 'Manoel Viana', 'Uruguaiana',
    ]),
    ('RO', 'Ji-Paraná', [
        "Alvorada D'Oeste", 'Costa Marques', 'Ji-Paraná',
        'Ouro Preto do Oeste', 'Presidente Médici',
        'São Miguel do Guaporé', 'Mirante da Serra', 'Nova União',
        'São Francisco do Guaporé', 'Seringueiras', 'Teixeirópolis',
        'Urupá', 'Vale do Paraíso',
    ]),
    ('RR', 'Rorainópolis', [
        'Caroebe', 'Rorainópolis', 'São João da Baliza', 'São Luiz',
    ]),
    ('SC', 'Curitibanos', [
        'Brunópolis', 'Curitibanos', 'Frei Rogério', 'Ponte Alta do Norte',
        'Santa Cecília', 'São Cristóvão do Sul',
    ]),
    ('SP', 'Itapeva', [
        'Apiaí', 'Barão de Antonina', 'Barra do Chapéu',
        'Bom Sucesso de Itararé', 'Buri', 'Capão Bonito', 'Guapiara',
        'Itaberá', 'Itaóca', 'Itapeva', 'Itapirapuã Paulista',
        'Itaporanga', 'Itararé', 'Nova Campina', 'Ribeira',
        'Ribeirão Branco', 'Ribeirão Grande', 'Riversul', 'Taquarivaí',
    ]),
    ('SE', 'Itabaiana', [
        'Areia Branca', 'Campo do Brito', 'Carira', 'Frei Paulo',
        'Itabaiana', 'Macambira', 'Malhador', 'Moita Bonita',
        'Nossa Senhora Aparecida', 'Pedra Mole', 'Pinhão',
        'Ribeirópolis', 'São Domingos', 'São Miguel do Aleixo',
    ]),
    ('TO', 'Araguaína', [
        'Ananás', 'Angico', 'Aragominas', 'Araguaína', 'Araguanã',
        'Arapoema', 'Babaçulândia', 'Barra do Ouro', 'Campos Lindos',
        'Carmolândia', 'Darcinópolis', 'Filadélfia', 'Goiatins',
        'Muricilândia', 'Nova Olinda', "Pau D'Arco", 'Piraquê',
        'Riachinho', 'Santa Fé do Araguaia', 'Wanderlândia', 'Xambioá',
    ]),
]


# Renomeações IBGE conhecidas (uf, nome_no_pdf) -> nome_no_banco.
ALIASES = {
    ('RN', 'Augusto Severo'): "Olho D'Água Do Borges",
}


# Tokens removidos no matching para tolerar variações de preposição
# (ex: "Muquém de São Francisco" no PDF vs "Muquém Do São Francisco" no banco).
_STOPWORDS = {'de', 'do', 'da', 'dos', 'das', 'e'}


def _norm(text):
    """Normaliza string para matching: lower, sem acentos, sem hífen/apóstrofo."""
    if not text:
        return ''
    nf = unicodedata.normalize('NFKD', str(text))
    nf = ''.join(c for c in nf if not unicodedata.combining(c))
    return (nf.lower()
              .replace("'", '')
              .replace('"', '')
              .replace('-', ' ')
              .strip())


def _norm_loose(text):
    """Como _norm, mas tira preposições para matching tolerante."""
    tokens = _norm(text).split()
    return ' '.join(t for t in tokens if t not in _STOPWORDS)


def _match(uf, nome):
    """Tenta achar o município por uf + name_muni com normalização tolerante."""
    nome_alvo = ALIASES.get((uf, nome), nome)
    alvo_norm = _norm(nome_alvo)
    alvo_loose = _norm_loose(nome_alvo)
    qs = list(Municipio.objects.filter(uf=uf))
    for m in qs:
        if _norm(m.name_muni) == alvo_norm:
            return m
    for m in qs:
        if _norm_loose(m.name_muni) == alvo_loose:
            return m
    return None


def _format_brl(value):
    """Formata float como string 'R$ 1.234,56' (padrão brasileiro)."""
    if value is None:
        return None
    return ('R$ ' + f'{float(value):,.2f}'
            .replace(',', 'X').replace('.', ',').replace('X', '.'))


def main():
    rows = []
    nao_encontrados = []

    for uf, ri, munis in GRUPOS:
        for nome in munis:
            m = _match(uf, nome)
            if m is None:
                nao_encontrados.append((uf, ri, nome))
                rows.append({
                    'uf': uf,
                    'ri': ri,
                    'municipio_pdf': nome,
                    'municipio_db': '—',
                    'rc_pc': None,
                    'rank_nacional_rc': None,
                    'quintil': '—',
                    'decil': '—',
                    'populacao': None,
                    'rank_nacional_pop': None,
                    'status': 'Não encontrado',
                })
                continue
            rows.append({
                'uf': uf,
                'ri': ri,
                'municipio_pdf': nome,
                'municipio_db': m.name_muni,
                'rc_pc': _format_brl(m.rc_24_pc),
                'rank_nacional_rc': m.rank_nacional,
                'quintil': m.quintil24 or '—',
                'decil': m.decil24 or '—',
                'populacao': m.populacao24,
                'rank_nacional_pop': m.populacao24_rank_nacional,
                'status': 'OK',
            })

    # ---- Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cidades Intermediadoras'

    headers = [
        'UF', 'Região Imediata (RI)', 'Município (PDF)', 'Município (DB)',
        'Receita Per Capita 2024',
        'Ranking Nacional Receita',
        'Quintil', 'Decil',
        'População 2024', 'Ranking Nacional População',
        'Status',
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='103758')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r in rows:
        ws.append([
            r['uf'], r['ri'], r['municipio_pdf'], r['municipio_db'],
            r['rc_pc'],
            r['rank_nacional_rc'],
            r['quintil'], r['decil'],
            r['populacao'], r['rank_nacional_pop'],
            r['status'],
        ])

    widths = [6, 30, 28, 28, 22, 20, 14, 14, 16, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'cidades_intermediadoras_ifem.xlsx')
    wb.save(out)
    print(f'Arquivo salvo: {out}')
    print(f'Total: {len(rows)} linhas | Não encontrados: {len(nao_encontrados)}')
    if nao_encontrados:
        print('\nMunicípios não encontrados (verificar grafia):')
        for uf, ri, nome in nao_encontrados:
            print(f'  {uf} | {ri} | {nome}')


if __name__ == '__main__':
    main()
